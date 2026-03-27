from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import HttpResponse
from django.utils import timezone
from datetime import timedelta, datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Apontamento
from .utils import is_owner

@login_required
@user_passes_test(is_owner)
def exportar_relatorio_excel(request):
    """
    Gera um relatório consolidado em Excel para conferência de folha e custos.
    """
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    queryset = Apontamento.objects.select_related(
        'projeto', 'colaborador', 'veiculo', 'centro_custo', 'codigo_cliente', 'registrado_por'
    ).prefetch_related('auxiliares_extras').all().order_by('data_apontamento', 'colaborador__nome_completo')
    
    if start_date_str and end_date_str:
        try:
            start = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            queryset = queryset.filter(data_apontamento__gte=start, data_apontamento__lte=end)
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatorio de Horas"

    headers = [
        "Data", "Dia Semana", "Colaborador", "Cargo", "Tipo", 
        "Local (Obra/Setor)", "Código de Obra", "Código Cliente", 
        "Veículo", "Placa", "Hora Início", "Hora Fim", "Total Horas", 
        "Plantão", "Dorme Fora", "Observações", "Registrado Por", 'Latitude', 'Longitude'
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    def get_duration_value(inicio, fim):
        if not inicio or not fim: return None
        
        dummy_date = date(2000, 1, 1)
        dt_inicio = datetime.combine(dummy_date, inicio)
        dt_fim = datetime.combine(dummy_date, fim)
        
        if dt_fim < dt_inicio:
            dt_fim += timedelta(days=1)

        return dt_fim - dt_inicio

    dias_semana_pt = {
        0: 'Segunda-feira', 1: 'Terça-feira', 2: 'Quarta-feira',
        3: 'Quinta-feira', 4: 'Sexta-feira', 5: 'Sábado', 6: 'Domingo'
    }

    for item in queryset:
        data_fmt = item.data_apontamento.strftime('%d/%m/%Y')
        dia_semana = dias_semana_pt[item.data_apontamento.weekday()]
        
        local_nome = ""
        col_codigo_obra = ""
        col_codigo_cliente = ""

        if item.local_execucao == 'INT':
            tipo = "OBRA"
            if item.projeto:
                local_nome = item.projeto.nome
                col_codigo_obra = item.projeto.codigo
            elif item.codigo_cliente:
                local_nome = item.codigo_cliente.nome
                col_codigo_cliente = item.codigo_cliente.codigo
        else:
            tipo = "FORA DO SETOR"
            local_nome = item.centro_custo.nome if item.centro_custo else "Atividade Externa"
            if item.projeto: col_codigo_obra = item.projeto.codigo
            elif item.codigo_cliente: col_codigo_cliente = item.codigo_cliente.codigo

        if col_codigo_obra and len(str(col_codigo_obra)) >= 5:
             if not col_codigo_cliente: col_codigo_cliente = str(col_codigo_obra)[1:5]
        elif col_codigo_obra:
             col_codigo_cliente = col_codigo_obra

        veiculo_nome_modelo = ""
        veiculo_placa_only = ""

        if item.veiculo:
            veiculo_nome_modelo = item.veiculo.descricao if item.veiculo.descricao else "Veículo da Frota"
            veiculo_placa_only = item.veiculo.placa
        elif item.veiculo_manual_modelo:
            veiculo_nome_modelo = item.veiculo_manual_modelo
            veiculo_placa_only = item.veiculo_manual_placa if item.veiculo_manual_placa else ""

        duracao_val = get_duration_value(item.hora_inicio, item.hora_termino)
        reg_por = item.registrado_por.username if item.registrado_por else "Sistema"

        plantao_str = "SIM" if item.em_plantao else "NÃO"
        dorme_fora_str = "SIM" if item.dorme_fora else "NÃO"

        row_principal = [
            data_fmt, dia_semana, item.colaborador.nome_completo, item.colaborador.cargo,
            tipo, local_nome, col_codigo_obra, col_codigo_cliente, 
            veiculo_nome_modelo, veiculo_placa_only, item.hora_inicio, item.hora_termino, 
            duracao_val,
            plantao_str, dorme_fora_str, 
            item.ocorrencias, reg_por,
            item.latitude, item.longitude
        ]
        ws.append(row_principal)

        cell_duration = ws.cell(row=ws.max_row, column=13)
        cell_duration.number_format = '[h]:mm:ss'

        auxiliares = []
        if item.auxiliar: auxiliares.append(item.auxiliar)
        auxiliares.extend(list(item.auxiliares_extras.all()))

        for aux in auxiliares:
            row_aux = [
                data_fmt, dia_semana, aux.nome_completo, aux.cargo,
                tipo, local_nome, col_codigo_obra, col_codigo_cliente, 
                "Carona", "", item.hora_inicio, item.hora_termino, 
                duracao_val,
                plantao_str, dorme_fora_str, 
                f"Auxiliar de: {item.colaborador.nome_completo}", reg_por,
                None, None
            ]
            ws.append(row_aux)
            
            cell_duration_aux = ws.cell(row=ws.max_row, column=13)
            cell_duration_aux.number_format = '[h]:mm:ss'

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        
        if column == 'P': 
            ws.column_dimensions[column].width = 50
            continue

        for cell in col:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length: max_length = val_len
            except: pass
            
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 40)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"Relatorio_Horas_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    wb.save(response)
    return response